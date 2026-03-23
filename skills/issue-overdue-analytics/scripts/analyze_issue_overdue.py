#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
问题列表超期分析统计脚本
按超期类型分类，并按处理人汇总统计
输出：
1. Excel文件 - 每个分类一个工作表 + 按处理人统计
2. 文本报告 - 打印汇总统计和最紧急问题列表，可直接复制发送
"""

import pandas as pd
import sys
import re
import os
from datetime import datetime


def extract_days(warning_text):
    """从超期预警提取天数"""
    numbers = re.findall(r'\d+\.?\d*', str(warning_text))
    if numbers:
        return float(numbers[0])
    return None


def format_issue_list(filtered, category):
    """格式化问题列表为文本"""
    lines = []
    
    if category == '签收超期':
        # 已经按紧急程度排序
        for _, row in filtered.iterrows():
            warning = str(row['超期预警'])
            title = row['问题标题']
            project = row['所属项目']
            handler = row['处理人']
            deadline = row['应完成时间']
            
            if '签收超期' in warning:
                days = extract_days(warning) or 0
                lines.append(f"1. [已超期]")
                lines.append(f"- 问题标题：{title}")
                lines.append(f"- 所属项目：{project}")
                lines.append(f"- 处理人：{handler}")
                lines.append(f"- 超期：{int(days*24)}小时")
                lines.append(f"- 应完成时间：{deadline}")
            elif '距离签收超期还有' in warning:
                days = extract_days(warning) or 0
                hours = int(days * 24)
                lines.append(f"1. [还有{hours}小时超期]")
                lines.append(f"- 问题标题：{title}")
                lines.append(f"- 所属项目：{project}")
                lines.append(f"- 处理人：{handler}")
                lines.append(f"- 应完成时间：{deadline}")
    
    elif category == '处理超期':
        for _, row in filtered.iterrows():
            title = row['问题标题']
            project = row['所属项目']
            handler = row['处理人']
            warning = str(row['超期预警'])
            deadline = row['应完成时间']
            lines.append(f"1. {title}")
            lines.append(f"- 所属项目：{project}")
            lines.append(f"- 处理人：{handler}")
            lines.append(f"- 超期预警：{warning}")
            lines.append(f"- 应完成时间：{deadline}")
    
    elif category == '剩余处理时间':
        # 已经按剩余天数排序
        for _, row in filtered.iterrows():
            warning = str(row['超期预警'])
            title = row['问题标题']
            handler = row['处理人']
            deadline = row['应完成时间']
            days = extract_days(warning) or 0
            lines.append(f"1. [剩余{days}天] {title}")
            lines.append(f"- 处理人：{handler}，应完成：{deadline}")
    
    elif category == '已处理待确认超期':
        for _, row in filtered.iterrows():
            title = row['问题标题']
            handler = row['处理人']
            warning = str(row['超期预警'])
            lines.append(f"1. {title}")
            lines.append(f"- 处理人：{handler}，超期预警：{warning}")
    
    return lines


def print_summary(categories, handler_stats, all_overdue):
    """打印汇总统计报告"""
    print("\n" + "="*60)
    print("📊 分类统计总览")
    print("分类\t数量")
    total = 0
    emoji_map = {
        '签收超期': '🔴',
        '处理超期': '🟠', 
        '剩余处理时间': '🟡',
        '已处理待确认超期': '🟢'
    }
    for name, df in categories.items():
        cnt = len(df)
        total += cnt
        emoji = emoji_map.get(name, '')
        print(f"{emoji} {name}\t{cnt} 条")
    print(f"合计\t{total} 条")
    
    # 打印详细列表
    for name, df in categories.items():
        if len(df) == 0:
            continue
        emoji = emoji_map.get(name, '')
        print(f"\n{emoji} {name}（按紧急程度排序）")
        lines = format_issue_list(df, name)
        for line in lines:
            print(line)
    
    # 按处理人统计
    print("\n📊 按处理人统计所有超期问题")
    print("处理人\t超期问题数量")
    handler_stats = handler_stats.sort_values('超期问题数量', ascending=False)
    for _, row in handler_stats.iterrows():
        print(f"{row['处理人']}\t{row['超期问题数量']}")
    
    # 找出最紧急的问题
    print("\n⚠️ 最紧急需要处理的：")
    # 已超期签收排前面
    urgent = []
    
    # 已超期签收（负号使排序正确，超期越久越靠前）
    if '签收超期' in categories:
        df = categories['签收超期']
        for _, row in df.iterrows():
            warning = str(row['超期预警'])
            if '签收超期' in warning:
                days = extract_days(warning) or 0
                hours = int(days * 24)
                urgent.append((-hours, f"- {row['处理人']}: 签收超期{hours}小时（严重超期） - {row['问题标题']}"))
            elif '距离签收超期还有' in warning:
                days = extract_days(warning) or 0
                hours = int(days * 24)
                if hours <= 24:
                    urgent.append((hours, f"- {row['处理人']}: {row['问题标题']}，还有{hours}小时超期，应完成{row['应完成时间']}"))
    
    # 处理超期也很紧急
    if '处理超期' in categories:
        df = categories['处理超期']
        for _, row in df.iterrows():
            days = extract_days(str(row['超期预警'])) or 0
            urgent.append((-days, f"- {row['处理人']}: {row['问题标题']}，处理超期{days}天"))
    
    # 按紧急程度排序
    urgent.sort(key=lambda x: x[0])
    for _, msg in urgent[:10]:  # 只显示前10个最紧急的
        print(msg)
    
    print("="*60)


def main():
    if len(sys.argv) < 2:
        print("Usage: python analyze_issue_overdue.py <input_excel> [output_excel]")
        print("Example: python analyze_issue_overdue.py 问题列表.xlsx 问题超期分类汇总.xlsx")
        sys.exit(1)
    
    input_file = sys.argv[1]
    if len(sys.argv) >= 3:
        output_file = sys.argv[2]
    else:
        # 默认输出文件名
        base_name = os.path.splitext(input_file)[0]
        output_file = f"{base_name}_分类汇总.xlsx"
    
    print(f"读取输入文件: {input_file}")
    df = pd.read_excel(input_file)
    print(f"共 {len(df)} 条问题记录")
    
    # 按关键词分类
    categories = {
        '签收超期': df['超期预警'].astype(str).str.contains('签收'),
        '处理超期': df['超期预警'].astype(str).str.contains('处理超期') & ~df['超期预警'].astype(str).str.contains('签收'),
        '剩余处理时间': df['超期预警'].astype(str).str.contains('剩余处理时间为'),
        '已处理待确认超期': df['超期预警'].astype(str).str.contains('已处理待确认超期'),
    }
    
    # 处理每个分类并保存筛选结果
    filtered_results = {}
    for cat_name, filter_mask in categories.items():
        filtered = df[filter_mask].copy()
        print(f"【{cat_name}】共 {len(filtered)} 条")
        
        # 排序
        if cat_name == '签收超期':
            # 提取紧急程度排序
            def get_sign_urgency(row):
                warning = str(row['超期预警'])
                numbers = re.findall(r'\d+\.?\d*', warning)
                if numbers:
                    num = float(numbers[0])
                    if '签收超期' in warning:
                        return -num  # 已超期，负数，超期越久越靠前
                    elif '距离签收超期还有' in warning:
                        return num  # 即将超期，时间越近越靠前
                return 999
            filtered['_urgency'] = filtered.apply(get_sign_urgency, axis=1)
            filtered = filtered.sort_values('_urgency')
            filtered = filtered.drop(columns=['_urgency'])
        
        elif cat_name == '剩余处理时间':
            # 按剩余天数排序
            def get_remaining_days(row):
                warning = str(row['超期预警'])
                numbers = re.findall(r'\d+\.?\d*', warning)
                if numbers:
                    return float(numbers[0])
                return 999
            filtered['_days'] = filtered.apply(get_remaining_days, axis=1)
            filtered = filtered.sort_values('_days')
            filtered = filtered.drop(columns=['_days'])
        
        filtered_results[cat_name] = filtered
    
    # 添加按处理人统计工作表
    all_overdue = pd.concat(filtered_results.values()).drop_duplicates()
    handler_stats = all_overdue.groupby('处理人').size().reset_index()
    handler_stats.columns = ['处理人', '超期问题数量']
    handler_stats = handler_stats.sort_values('超期问题数量', ascending=False)
    
    # 创建汇总统计DataFrame
    summary_data = []
    emoji_map = {
        '签收超期': '🔴',
        '处理超期': '🟠', 
        '剩余处理时间': '🟡',
        '已处理待确认超期': '🟢'
    }
    total = 0
    for name, df in filtered_results.items():
        cnt = len(df)
        total += cnt
        summary_data.append({
            '分类': f"{emoji_map.get(name, '')} {name}",
            '数量': f"{cnt} 条"
        })
    summary_data.append({
        '分类': '合计',
        '数量': f"{total} 条"
    })
    summary_df = pd.DataFrame(summary_data)
    
    # 输出文本汇总报告
    # 直接生成文本报告，不修改内置print
    def generate_text_report():
        lines = []
        lines.append("="*60)
        lines.append("📊 分类统计总览")
        lines.append("分类\t数量")
        total = 0
        for name, df in filtered_results.items():
            cnt = len(df)
            total += cnt
            emoji = emoji_map.get(name, '')
            lines.append(f"{emoji} {name}\t{cnt} 条")
        lines.append(f"合计\t{total} 条")
        
        # 打印详细列表
        for name, df in filtered_results.items():
            if len(df) == 0:
                continue
            emoji = emoji_map.get(name, '')
            lines.append(f"\n{emoji} {name}（按紧急程度排序）")
            issue_lines = format_issue_list(df, name)
            lines.extend(issue_lines)
        
        # 按处理人统计
        lines.append("\n📊 按处理人统计所有超期问题")
        lines.append("处理人\t超期问题数量")
        for _, row in handler_stats.iterrows():
            lines.append(f"{row['处理人']}\t{row['超期问题数量']}")
        
        # 找出最紧急的问题
        lines.append("\n⚠️ 最紧急需要处理的：")
        # 已超期签收排前面
        urgent = []
        
        # 已超期签收（负号使排序正确，超期越久越靠前）
        if '签收超期' in filtered_results:
            df = filtered_results['签收超期']
            for _, row in df.iterrows():
                warning = str(row['超期预警'])
                if '签收超期' in warning:
                    days = extract_days(warning) or 0
                    hours = int(days * 24)
                    urgent.append((-hours, f"- {row['处理人']}: 签收超期{hours}小时（严重超期） - {row['问题标题']}"))
                elif '距离签收超期还有' in warning:
                    days = extract_days(warning) or 0
                    hours = int(days * 24)
                    if hours <= 24:
                        urgent.append((hours, f"- {row['处理人']}: {row['问题标题']}，还有{hours}小时超期，应完成{row['应完成时间']}"))
        
        # 处理超期也很紧急
        if '处理超期' in filtered_results:
            df = filtered_results['处理超期']
            for _, row in df.iterrows():
                days = extract_days(str(row['超期预警'])) or 0
                urgent.append((-days, f"- {row['处理人']}: {row['问题标题']}，处理超期{days}天"))
        
        # 按紧急程度排序
        urgent.sort(key=lambda x: x[0])
        for _, msg in urgent[:10]:  # 只显示前10个最紧急的
            lines.append(msg)
        
        lines.append("="*60)
        return '\n'.join(lines)
    
    report_text = generate_text_report()
    print(report_text)
    
    # 保存Excel - 在分类汇总工作表中包含汇总统计和所有明细
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment
    
    wb = Workbook()
    ws = wb.active
    ws.title = '分类汇总'
    
    # 写入汇总统计
    row_num = 1
    ws.cell(row=row_num, column=1, value='分类统计总览')
    ws.cell(row=row_num, column=1).font = Font(bold=True, size=14)
    row_num += 1
    
    ws.cell(row=row_num, column=1, value='分类')
    ws.cell(row=row_num, column=2, value='数量')
    ws.cell(row=row_num, column=1).font = Font(bold=True)
    ws.cell(row=row_num, column=2).font = Font(bold=True)
    row_num += 1
    
    emoji_map = {
        '签收超期': '🔴',
        '处理超期': '🟠', 
        '剩余处理时间': '🟡',
        '已处理待确认超期': '🟢'
    }
    total = 0
    for name, df in filtered_results.items():
        cnt = len(df)
        total += cnt
        ws.cell(row=row_num, column=1, value=f"{emoji_map.get(name, '')} {name}")
        ws.cell(row=row_num, column=2, value=f"{cnt} 条")
        row_num += 1
    
    ws.cell(row=row_num, column=1, value='合计')
    ws.cell(row=row_num, column=2, value=f"{total} 条")
    ws.cell(row=row_num, column=1).font = Font(bold=True)
    ws.cell(row=row_num, column=2).font = Font(bold=True)
    row_num += 2  # 空两行
    
    # 写入每个分类的明细
    for name, df in filtered_results.items():
        if len(df) == 0:
            continue
        
        # 分类标题
        ws.cell(row=row_num, column=1, value=f"{emoji_map.get(name, '')} {name}（按紧急程度排序）")
        ws.cell(row=row_num, column=1).font = Font(bold=True)
        row_num += 1
        
        # 表头 - 统一顺序：超期信息、应完成时间、处理人、问题标题、所属项目
        headers = ['序号', '超期信息', '应完成时间', '处理人', '问题标题', '所属项目']
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_num, column=col_idx, value=header)
            cell.font = Font(bold=True)
        row_num += 1
        
        # 写入明细数据
        for idx, (_, row) in enumerate(df.iterrows(), 1):
            warning = str(row['超期预警'])
            
            # 提取超期信息
            if name == '签收超期':
                days = extract_days(warning) or 0
                hours = int(days * 24)
                if '签收超期' in warning:
                    overdue_info = f'已超期 {hours}小时'
                elif '距离签收超期还有' in warning:
                    overdue_info = f'还有 {hours}小时 超期'
                else:
                    overdue_info = warning
            elif name == '处理超期':
                overdue_info = warning
            elif name == '剩余处理时间':
                days = extract_days(warning) or 0
                overdue_info = f'剩余 {days}天'
            elif name == '已处理待确认超期':
                overdue_info = warning
            
            # 统一顺序：超期信息、应完成时间、处理人、问题标题、所属项目
            data = [
                idx,
                overdue_info,
                str(row['应完成时间']),
                row['处理人'],
                row['问题标题'],
                row['所属项目']
            ]
            
            for col_idx, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col_idx, value=value)
            row_num += 1
        
        row_num += 1  # 空一行
    
    # 按处理人统计
    ws.cell(row=row_num, column=1, value='📊 按处理人统计所有超期问题')
    ws.cell(row=row_num, column=1).font = Font(bold=True, size=12)
    row_num += 1
    
    ws.cell(row=row_num, column=1, value='处理人')
    ws.cell(row=row_num, column=2, value='超期问题数量')
    ws.cell(row=row_num, column=1).font = Font(bold=True)
    ws.cell(row=row_num, column=2).font = Font(bold=True)
    row_num += 1
    
    for _, row in handler_stats.iterrows():
        ws.cell(row=row_num, column=1, value=row['处理人'])
        ws.cell(row=row_num, column=2, value=row['超期问题数量'])
        row_num += 1
    
    row_num += 1
    
    # 最紧急需要处理的
    ws.cell(row=row_num, column=1, value='⚠️ 最紧急需要处理的')
    ws.cell(row=row_num, column=1).font = Font(bold=True)
    row_num += 1
    
    # 找出最紧急的问题
    urgent = []
    # 已超期签收排前面
    if '签收超期' in filtered_results:
        df = filtered_results['签收超期']
        for _, row in df.iterrows():
            warning = str(row['超期预警'])
            if '签收超期' in warning:
                days = extract_days(warning) or 0
                hours = int(days * 24)
                urgent.append((-hours, f"{row['处理人']}: 签收超期{hours}小时（严重超期） - {row['问题标题']}"))
            elif '距离签收超期还有' in warning:
                days = extract_days(warning) or 0
                hours = int(days * 24)
                if hours <= 24:
                    urgent.append((hours, f"{row['处理人']}: {row['问题标题']}，还有{hours}小时超期，应完成{row['应完成时间']}"))
    
    # 处理超期也很紧急
    if '处理超期' in filtered_results:
        df = filtered_results['处理超期']
        for _, row in df.iterrows():
            days = extract_days(str(row['超期预警'])) or 0
            urgent.append((-days, f"{row['处理人']}: {row['问题标题']}，处理超期{days}天"))
    
    # 按紧急程度排序
    urgent.sort(key=lambda x: x[0])
    for _, msg in urgent[:10]:  # 只显示前10个最紧急的
        ws.cell(row=row_num, column=1, value=msg)
        row_num += 1
    
    # 调整列宽
    ws.column_dimensions['A'].width = 8    # 序号
    ws.column_dimensions['B'].width = 20   # 超期信息
    ws.column_dimensions['C'].width = 25   # 应完成时间
    ws.column_dimensions['D'].width = 30   # 处理人
    ws.column_dimensions['E'].width = 50   # 问题标题
    ws.column_dimensions['F'].width = 40   # 所属项目
    
    # 保存文件 - 只保留一个工作表，不添加独立的按处理人统计sheet
    wb.save(output_file)
    print(f"\n✨ 分类汇总完成")
    print(f"📊 Excel输出: {output_file}")
    total_unique = len(all_overdue)
    print(f"📝 总计去重后: {total_unique} 条超期问题")
    
    print(f"\n✨ 分类汇总完成")
    print(f"📊 Excel输出: {output_file}")
    total_unique = len(all_overdue)
    print(f"📝 总计去重后: {total_unique} 条超期问题")


if __name__ == '__main__':
    main()
